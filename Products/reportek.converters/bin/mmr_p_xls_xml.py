# The contents of this file are subject to the Mozilla Public
# License Version 1.1 (the "License"); you may not use this file
# except in compliance with the License. You may obtain a copy of
# the License at http://www.mozilla.org/MPL/
#
# Software distributed under the License is distributed on an "AS
# IS" basis, WITHOUT WARRANTY OF ANY KIND, either express or
# implied. See the License for the specific language governing
# rights and limitations under the License.
#
# The Initial Owner of the Original Code is European Environment
# Agency (EEA). Portions created by Finsiel Romania are
# Copyright (C) European Environment Agency. All
# Rights Reserved.
#
# Authors:
# Olimpiu Rob - Eau de Web Romania

__doc__ = """
    MMR Projections XLS to XML converter module.
"""

from decimal import Decimal
from lxml import etree
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from utils import utOpen
import argparse
import requests
import sys
from cachetools import cached, TTLCache
cache = TTLCache(maxsize=100, ttl=3600)
SCHEMA = "http://dd.eionet.europa.eu/schemas/mmr-projections/projections-Article23table1v6.XSD"


@cached(cache)
def get_schema(url):
    try:
        response = requests.get(url, verify=False)
    except Exception as e:
        return None
    if response.status_code != requests.codes.ok:
        return None

    tree = etree.fromstring(response.content)
    return tree


def get_cat_tag(name, nsmap=None):
    tree = get_schema(SCHEMA)
    if not nsmap:
        nsmap = {'xsd': 'http://www.w3.org/2001/XMLSchema'}
    xpath = ".//xsd:enumeration[xsd:annotation/xsd:documentation='{}']".format(name)
    elems = tree.xpath(xpath, namespaces=nsmap)
    if len(elems) == 1:
        return elems[0].attrib['value']


def get_name_boundaries(wb, name):
    """ return a tuple of boundaries: (min_col, min_row, max_col, max_row)"""
    val = wb.defined_names[name].value.split('!')[-1]
    return range_boundaries(val)


def mmr_p_xls_to_xml(xls):
    input_xls = utOpen(xls)
    wb = load_workbook(input_xls, data_only=True)
    # We only convert the first sheet
    ws = wb.worksheets[0]
    # Grab the ranges for the tags
    val_coords = get_name_boundaries(wb, 'Values')
    years_coords = get_name_boundaries(wb, 'AYears')
    gu_coords = get_name_boundaries(wb, 'AGasUnits')
    # Set up the QNAME
    xsi = "http://www.w3.org/2001/XMLSchema-instance"
    attr_qname = etree.QName("http://www.w3.org/2001/XMLSchema-instance",
                             "noNamespaceSchemaLocation")
    nsmap = {'xsi': 'http://www.w3.org/2001/XMLSchema-instance'}
    root = etree.Element("ProjectionsTable1",
                         {attr_qname: SCHEMA},
                         nsmap=nsmap)
    isy = etree.Element("Inventory_Submission_year")
    isy.text = str(ws.cell(row=14, column=2).value)
    ms = etree.Element("MS")
    ms.text = str(ws.cell(row=15, column=2).value)
    root.append(isy)
    root.append(ms)
    for row in ws.iter_rows(min_row=val_coords[1]):
        catv = str(row[0].value).strip()
        scenariov = str(row[1].value)
        for idx in range(val_coords[0]-1, val_coords[2]):
            value = row[idx].value
            if value is not None:
                rowxml = etree.Element("Row")
                root.append(rowxml)
                cur_col = row[idx].col_idx
                cat = etree.Element("Category__1_3")
                cat.text = get_cat_tag(catv)
                rowxml.append(cat)
                year = etree.Element("Year")
                # Grab the year tag for the current column
                year.text = str(ws.cell(row=years_coords[1], column=cur_col).value)
                rowxml.append(year)
                scenario = etree.Element("Scenario")
                scenario.text = scenariov
                rowxml.append(scenario)
                gu = etree.Element("Gas___Units")
                # Grab the gas unit tag for the current column
                gu.text = ws.cell(row=gu_coords[1], column=cur_col).value
                rowxml.append(gu)
                val = etree.Element("Value")
                nk = etree.Element("NK")
                if isinstance(value, str) or isinstance(value, unicode):
                    nk.text = value
                else:
                    val.text = repr(value)
                rowxml.append(nk)
                rowxml.append(val)

    return etree.tostring(root, xml_declaration=True,
                          encoding='UTF-8', pretty_print=True)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument(
        'xls',
        metavar='xls',
        help='path to xls file')
    try:
        args = parser.parse_args()
    except:
        args = None

    if not args or not args.xls:
        print __doc__
        print "For help use --help"
    else:
        if sys.platform == "win32":
            msvcrt.setmode(sys.stdout.fileno(), os.O_BINARY)
        sys.stdout.write(mmr_p_xls_to_xml(xls=args.xls))
